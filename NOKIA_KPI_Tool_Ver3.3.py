# -*- coding: utf-8 -*-
"""
Created on Thu Mar 10 18:33:56 2022

@author: vatsalgu
"""


tech_names={
    "G":"GSM",
    "P":"U1900",
    "U":"U2100",
    "E":"L600",
    "D":"L700",
    "B":"L1900",
    "L":"L2100",
    "T":"L2500",
    "F":"AWS3",
    "Z":"IOT",
    "Y":"IOT",
    "K":"NR600",
    "A":"NR2500",
    "N":"N260",
    "M":"MMW"
}

sector_names={
    "1":"A",
    "2":"B",
    "3":"G",
    "4":"D",
    "5":"E",
    "6":"Z",
    "7":"7th sector",
    "8":"8th sector",
    "9":"9th sector",
    "A":"A",
    "B":"B",
    "C":"G",
    "D":"D",
    "E":"E",
    "F":"Z",
}

carrier_names={
    "1":"_1C",
    "2":"_2C",
    "3":"_3C"
}

duration_meaning={
    'KPI Hour1 Validated':'1',
    'KPI Hour2 Validated':'2',
    'KPI Hour4 Validated':'4',
    'KPI Hour8 Validated':'8',
    'KPI Next Day Validated':'24',
    'KPI Next 2 Days Validated':'48'

}

def update_cell_value(cell_val):
    if cell_val.find('GPA')!=-1:
        val = list(cell_val)
        return "GSM_1C (%s)" % (sector_names[val[-5]])
    else:
        val = list(cell_val)
        return "%s%s (%s)" % (tech_names[val[0]], carrier_names[val[-1]], sector_names[val[-2]])

def update_cell_value_ms(cell_val): #minus sector
    if cell_val.find('GPA')!=-1:
        val = list(cell_val)
        orig_return = "GSM_1C (%s)" % (sector_names[val[-5]])
    else:
        val = list(cell_val)
        orig_return = "%s%s (%s)" % (tech_names[val[0]], carrier_names[val[-1]], sector_names[val[-2]])
    # return orig_return
    if orig_return[0:3]=='GSM':
        return 'GSM'
    sectors_allowed = ['L1900', 'L2500', 'U1900', 'U2100', 'L2100', 'NR2500']
    if orig_return.split('_')[0] in sectors_allowed:
        return ' '.join(orig_return.split(' ')[:1])
    else:
        return ' '.join(orig_return.split('_')[:1])

def formatpara(mainstring):
    #mainstring='E-RAB Drop Rate Active Users GCR_L16 degraded at L19 2C (B)\nLTE Data Access Failure Rate GCR degraded at L19 2C (B), L19 2C (G)\nVoLTE Access Failure Rate degraded at L19 2C (A), L19 2C (B), AWS3 1C (A), L19 2C (G)\nRACH setup failure rate degraded at L19 1C (G), L19 2C (G)'
    if mainstring=='' or mainstring==None or mainstring=='\n':
        return mainstring
    
    string=mainstring.split('\n')
    finalli=[]
    for k in range(0,len(string)):    
        start_string=string[k][:string[k].index(' at ')+4]
        li=string[k][string[k].index(' at ')+3::].strip().split(', ')
        li = sorted(li, key = lambda x: x.split('_')[1].split()[0])
        newli=[]
        complete=''
        singleli=''
        
        groupli = [list(i) for j, i in groupby(li,lambda a: a.partition('C ')[0])]
        for i in range (0,len(groupli)):
            if len(groupli[i])==1:
                #newli=[]
                newli.append(''.join(groupli[i]))
                if len(newli)==1:
                    singleli=''.join(newli)
                else:
                    singleli=', '.join(newli)
            else:
                endsub=''
                endsub=groupli[i][0][groupli[i][0].index('('):groupli[i][0].index(')')]
                for j in range (1,len(groupli[i])):
                    startsub=groupli[i][j][:groupli[i][j].index('(')]
                    endsub+=groupli[i][j][groupli[i][j].index('(')+1:groupli[i][j].index(')')]
                endsub+=')'
                if len(groupli)==1 or i==0:
                    complete=startsub+endsub
                else:
                    complete+=', '+startsub+endsub
        if i==len(groupli)-1:
            if singleli=='':
                finalli.append(start_string+complete)
            elif complete=='':
                finalli.append(start_string+singleli)
            else:
                finalli.append(start_string+singleli+', '+complete)
            
        mainstring_res='1. '+finalli[0]
        for i in range(1,len(finalli)):
            mainstring_res+='\n'+str(i+1)+'. '+finalli[i]
        
    return mainstring_res

def create_all_antennas(df, site_totaltechs_list): #For total
    grouped = df.groupby('SITE')
    groups = list(grouped.groups.keys())

    for group in groups:
        existing = site_totaltechs_list.get(group, [])
        site_totaltechs_list[group] = list(set(existing + grouped.get_group(group).CELL.apply(lambda x : update_cell_value_ms(x)).tolist())) 
        #site_totaltechs_list[group] = list(set(grouped.get_group(group).CELL.apply(lambda x : update_cell_value_ms(x)).tolist()))

    return site_totaltechs_list

def generate_lte_remarks(site_list,fdataset,col_dict_df):
    final_output = {}
        
    main_output = {}
    for site in site_list:
        main_output[site] = []
        final_output[site] = []
        for row in fdataset.iloc:
            if row[col_dict_df["SITE"]] == site:
                if row[col_dict_df["Cell Availability, excluding blocked by user state"]] == 0:
                    val = update_cell_value(row[col_dict_df["CELL"]])
                    main_output[site].append(val)
        if len(main_output[site]) > 0:
            main_output[site] = ["0 Cell availability rate at %s" % str(", ".join(main_output[site]))]
        final_output[site].append(main_output[site])
    
    main_output = {}
    for site in site_list:
        main_output[site] = []
        final_output[site] = []
        for row in fdataset.iloc:
            if row[col_dict_df["SITE"]] == site:
                if row[col_dict_df["Cell Availability, excluding blocked by user state"]] > 0 and row[col_dict_df["Cell Availability, excluding blocked by user state"]] < 98:
                    val = update_cell_value(row[col_dict_df["CELL"]])
                    main_output[site].append(val)
        if len(main_output[site]) > 0:
            main_output[site] = ["Low Cell availability rate at %s" % str(", ".join(main_output[site]))]
        final_output[site].append(main_output[site]) 
        
    main_output = {}    
    for site in site_list:
        main_output[site] = []
        for row in fdataset.iloc:
            if row[col_dict_df["SITE"]] == site:
                if row[col_dict_df["E-RAB Drop Rate Active Users(GCR)_L16"]] > 2:
                    val = update_cell_value(row[col_dict_df["CELL"]])
                    main_output[site].append(val)
        if len(main_output[site]) > 0:
            main_output[site] = ["High E-RAB Drop Rate at %s" % str(", ".join(main_output[site]))]
        final_output[site].append(main_output[site]) 
        
    main_output = {}
    for site in site_list:
        main_output[site] = []
        for row in fdataset.iloc:
            if row[col_dict_df["SITE"]] == site:
                if row[col_dict_df["LTE Data Access Failure Rate GCR"]] > 2:
                    val = update_cell_value(row[col_dict_df["CELL"]])
                    main_output[site].append(val)
        if len(main_output[site]) > 0:
            main_output[site] = ["High LTE Data Access Failure Rate at %s" % str(", ".join(main_output[site]))]
        final_output[site].append(main_output[site])
    
    main_output = {}
    for site in site_list:
        main_output[site] = []
        for row in fdataset.iloc:
            if row[col_dict_df["SITE"]] == site:
                if (row[col_dict_df["VoLTE Access Failures"]] > 10 and row[col_dict_df["VoLTE Access Failure Rate (GCR)"]] > 2):
                    val = update_cell_value(row[col_dict_df["CELL"]])
                    main_output[site].append(val)
        if len(main_output[site]) > 0:
            main_output[site] = ["High VoLTE Access Failure Rate at %s" % str(", ".join(main_output[site]))]
        final_output[site].append(main_output[site])
    
    main_output = {}
    for site in site_list:
        main_output[site] = []
        for row in fdataset.iloc:
            if row[col_dict_df["SITE"]] == site:
                if (row[col_dict_df["Total VoLTE Drops L16"]] > 10 and row[col_dict_df["VoLTE Drop Rate (GCR) L16"]] > 2):
                    val = update_cell_value(row[col_dict_df["CELL"]])
                    main_output[site].append(val)
        if len(main_output[site]) > 0:
            main_output[site] = ["High VoLTE Drop Rate at %s" % str(", ".join(main_output[site]))]
        final_output[site].append(main_output[site])
    
    main_output = {}
    for site in site_list:
        main_output[site] = []
        for row in fdataset.iloc:
            if row[col_dict_df["SITE"]] == site:
                if row[col_dict_df["DL Traffic Volume (MB)"]] == 0:
                    val = update_cell_value(row[col_dict_df["CELL"]])
                    main_output[site].append(val)
        if len(main_output[site]) > 0:
            main_output[site] = ["0 DL Traffic Volume MB at %s" % str(", ".join(main_output[site]))]
        final_output[site].append(main_output[site])
    
    main_output = {}
    for site in site_list:
        main_output[site] = []
        for row in fdataset.iloc:
            if row[col_dict_df["SITE"]] == site:
                if row[col_dict_df["VoLTE Erlangs"]] == 0:
                    val = update_cell_value(row[col_dict_df["CELL"]])
                    main_output[site].append(val)
        if len(main_output[site]) > 0:
            main_output[site] = ["0 VoLTE Erlangs at %s" % str(", ".join(main_output[site]))]
        final_output[site].append(main_output[site])
            
#Can be neglected for now -->
    
        # main_output = {}    
        # for site in site_list:
        #     main_output[site] = []
        #     for row in dataset.iloc:
        #         if row[4] == site:
        #             if row[18] > -3:
        #                 val = update_cell_value(row[6])
        #                 main_output[site].append(val)
        #     if len(main_output[site]) > 0:
        #                 main_output[site] = ["Average RTWP Rx Antenna 1 degraded at %s" % str(", ".join(main_output[site]))]
        #     final_output[site].append(main_output[site])
        
        # main_output = {} 
        # for site in site_list:
        #     main_output[site] = []
        #     for row in dataset.iloc:
        #         if row[4] == site:
        #             if row[20] > -3:
        #                 val = update_cell_value(row[6])
        #                 main_output[site].append(val)
        #     if len(main_output[site]) > 0:
        #                 main_output[site] = ["Average RTWP Rx Antenna 2 degraded at %s" % str(", ".join(main_output[site]))]
        #     final_output[site].append(main_output[site])
        
        # main_output = {}
        # for site in site_list:
        #     main_output[site] = []
        #     for row in dataset.iloc:
        #         if row[4] == site:
        #             if row[19] > -3:
        #                 val = update_cell_value(row[6])
        #                 main_output[site].append(val)
        #     if len(main_output[site]) > 0:
        #                 main_output[site] = ["Average RTWP Rx Antenna 3 degraded at %s" % str(", ".join(main_output[site]))]
        #     final_output[site].append(main_output[site])
        
        
        # main_output = {}
        # for site in site_list:
        #     main_output[site] = []
        #     for row in dataset.iloc:
        #         if row[4] == site:
        #             if row[21] > -3:
        #                 val = update_cell_value(row[6])
        #                 main_output[site].append(val)
        #     if len(main_output[site]) > 0:
        #                 main_output[site] = ["Average RTWP Rx Antenna 4 degraded at %s" % str(", ".join(main_output[site]))]
        #     final_output[site].append(main_output[site])
    
#<-- Can be neglected for now.

    main_output = {}
    for site in site_list:
        main_output[site] = []
        for row in fdataset.iloc:
            if row[col_dict_df["SITE"]] == site:
                if row[col_dict_df["Inter-Frequency HO Success Rate"]] == 0:
                    val = update_cell_value(row[col_dict_df["CELL"]])
                    main_output[site].append(val)
        if len(main_output[site]) > 0:
            main_output[site] = ["0 Inter Frequency HO Success Rate at %s" % str(", ".join(main_output[site]))]
        final_output[site].append(main_output[site])
        
    main_output = {}
    for site in site_list:
        main_output[site] = []
        for row in fdataset.iloc:
            if row[col_dict_df["SITE"]] == site:
                if row[col_dict_df["Intra-Frequency HO Success Rate"]] == 0:
                    val = update_cell_value(row[col_dict_df["CELL"]])
                    main_output[site].append(val)
        if len(main_output[site]) > 0:
            main_output[site] = ["0 Intra Frequency HO Success Rate at %s" % str(", ".join(main_output[site]))]
        final_output[site].append(main_output[site])
    
    main_output = {}
    for site in site_list:
        main_output[site] = []
        for row in fdataset.iloc:
            if row[col_dict_df["SITE"]] == site:
                if row[col_dict_df["RACH setup failure rate"]] > 20:
                    val = update_cell_value(row[col_dict_df["CELL"]])
                    main_output[site].append(val)
        if len(main_output[site]) > 0:
            main_output[site] = ["High RACH setup failure rate at %s" % str(", ".join(main_output[site]))]
        final_output[site].append(main_output[site])
    
    #main_output = {}
    #for site in site_list:
    #    main_output[site] = []
    #    for row in fdataset.iloc:
    #        if row[col_dict_df["SITE"]] == site:
    #            if row[col_dict_df["RRC_CONNECTED_UE_AVG_M8051C55_"]] == 0:
    #                val = update_cell_value(row[col_dict_df["CELL"]])
    #                main_output[site].append(val)
    #    if len(main_output[site]) > 0:
    #        main_output[site] = ["0 RRC CONNECTED UE_AVG at %s" % str(", ".join(main_output[site]))]
    #    final_output[site].append(main_output[site])
        
    main_output = {}
    for site in site_list:
        main_output[site] = []
        for row in fdataset.iloc:
            if row[col_dict_df["SITE"]] == site:
                if row[col_dict_df["RRC_CONNECTED_UE_MAX_M8051C56_"]] == 0:
                    val = update_cell_value(row[col_dict_df["CELL"]])
                    main_output[site].append(val)
        if len(main_output[site]) > 0:
            main_output[site] = ["0 RRC CONNECTED UE_MAX at %s" % str(", ".join(main_output[site]))]
        final_output[site].append(main_output[site])
        
    main_output = {}
    for site in site_list:
        main_output[site] = []
        for row in fdataset.iloc:
            if row[col_dict_df["SITE"]] == site:
                if row[col_dict_df["UL Traffic Volume (MB)"]] == 0:
                    val = update_cell_value(row[col_dict_df["CELL"]])
                    main_output[site].append(val)
        if len(main_output[site]) > 0:
            main_output[site] = ["0 UL Traffic Volume MB at %s" % str(", ".join(main_output[site]))]
        final_output[site].append(main_output[site])
        
        
    for key, value in final_output.items():
        output = []
        for line in value:
            output += line
        final_output[key] = ["\n".join(output)]

    for k,v in final_output.items():
        final_output[k]=formatpara(''.join(final_output[k]))
        
    return final_output


def generate_5g_remarks(site_list,fdataset,col_dict_df):
    final_output = {}
        
    main_output = {}
    for site in site_list:
        main_output[site] = []
        final_output[site] = []
        for row in fdataset.iloc:
            if row[col_dict_df["SITE"]] == site:
                if row[col_dict_df["NR DL Traffic Vol - MAC (MB)"]] == 0:
                    val = update_cell_value(row[col_dict_df["CELL"]])
                    main_output[site].append(val)
        if len(main_output[site]) > 0:
            main_output[site] = ["0 NR DL Traffic Vol-MAC (MB) at %s" % str(", ".join(main_output[site]))]
        final_output[site].append(main_output[site]) 
        
    main_output = {}    
    for site in site_list:
        main_output[site] = []
        for row in fdataset.iloc:
            if row[col_dict_df["SITE"]] == site:
                if row[col_dict_df["NR UL Traffic Vol - MAC (MB)"]] == 0:
                    val = update_cell_value(row[col_dict_df["CELL"]])
                    main_output[site].append(val)
        if len(main_output[site]) > 0:
            main_output[site] = ["0 NR UL Traffic Vol-MAC (MB) at %s" % str(", ".join(main_output[site]))]
        final_output[site].append(main_output[site]) 
        
    main_output = {}
    for site in site_list:
        main_output[site] = []
        for row in fdataset.iloc:
            if row[col_dict_df["SITE"]] == site:
                if row[col_dict_df["OEM NR ACC NSA Accessibility (NR_5020d)"]] == 0:
                    val = update_cell_value(row[col_dict_df["CELL"]])
                    main_output[site].append(val)
        if len(main_output[site]) > 0:
            main_output[site] = ["0 NR NSA Accessibility at %s" % str(", ".join(main_output[site]))]
        final_output[site].append(main_output[site])
    
    main_output = {}
    for site in site_list:
        main_output[site] = []
        for row in fdataset.iloc:
            if row[col_dict_df["SITE"]] == site:
                if row[col_dict_df["OEM NR ACC NSA Accessibility (NR_5020d)"]] > 0 and row[col_dict_df["OEM NR ACC NSA Accessibility (NR_5020d)"]] < 98:
                    val = update_cell_value(row[col_dict_df["CELL"]])
                    main_output[site].append(val)
        if len(main_output[site]) > 0:
            main_output[site] = ["Low NR NSA Accessibility at %s" % str(", ".join(main_output[site]))]
        final_output[site].append(main_output[site])
        
    main_output = {}
    for site in site_list:
        main_output[site] = []
        for row in fdataset.iloc:
            if row[col_dict_df["SITE"]] == site:
                if row[col_dict_df["OEM NR AVL Cell Availability (NR_5150a)"]] == 0:
                    val = update_cell_value(row[col_dict_df["CELL"]])
                    main_output[site].append(val)
        if len(main_output[site]) > 0:
            main_output[site] = ["0 NR Cell Availability at %s" % str(", ".join(main_output[site]))]
        final_output[site].append(main_output[site])
    
    main_output = {}
    for site in site_list:
        main_output[site] = []
        for row in fdataset.iloc:
            if row[col_dict_df["SITE"]] == site:
                if row[col_dict_df["OEM NR AVL Cell Availability (NR_5150a)"]] > 0 and row[col_dict_df["OEM NR AVL Cell Availability (NR_5150a)"]] < 98:
                    val = update_cell_value(row[col_dict_df["CELL"]])
                    main_output[site].append(val)
        if len(main_output[site]) > 0:
            main_output[site] = ["Low NR Cell Availability at %s" % str(", ".join(main_output[site]))]
        final_output[site].append(main_output[site])
        
        
    for key, value in final_output.items():
        output = []
        for line in value:
            output += line
        final_output[key] = ["\n".join(output)]
        
    for k,v in final_output.items():
        final_output[k]=formatpara(''.join(final_output[k]))
    
    return final_output


def generate_gsm_remarks(site_list,fdataset,col_dict_df):
    final_output = {}
        
    main_output = {}
    for site in site_list:
        main_output[site] = []
        final_output[site] = []
        for row in fdataset.iloc:
            if row[col_dict_df["SITE"]] == site:
                if row[col_dict_df["% BCCH Availability"]] == 0:
                    val = update_cell_value(row[col_dict_df["CELL"]])
                    main_output[site].append(val)
        if len(main_output[site]) > 0:
            main_output[site] = ["0 %"+" BCCH Availability at %s" % str(", ".join(main_output[site]))]
        final_output[site].append(main_output[site]) 
        
    main_output = {}
    for site in site_list:
        main_output[site] = []
        for row in fdataset.iloc:
            if row[col_dict_df["SITE"]] == site:
                if row[col_dict_df["% BCCH Availability"]] > 0 and row[col_dict_df["% BCCH Availability"]] < 98:
                    val = update_cell_value(row[col_dict_df["CELL"]])
                    main_output[site].append(val)
        if len(main_output[site]) > 0:
            main_output[site] = ["Low %"+" BCCH Availability at %s" % str(", ".join(main_output[site]))]
        final_output[site].append(main_output[site])
    
    main_output = {}    
    for site in site_list:
        main_output[site] = []
        for row in fdataset.iloc:
            if row[col_dict_df["SITE"]] == site:
                if row[col_dict_df["% Blocked (calls and Inc HO)"]] > 2:
                    val = update_cell_value(row[col_dict_df["CELL"]])
                    main_output[site].append(val)
        if len(main_output[site]) > 0:
            main_output[site] = ["High %"+" Blocked (calls and Inc HO) at %s" % str(", ".join(main_output[site]))]
        final_output[site].append(main_output[site]) 
        
    main_output = {}
    for site in site_list:
        main_output[site] = []
        for row in fdataset.iloc:
            if row[col_dict_df["SITE"]] == site:
                if row[col_dict_df["SDCCH Blocking"]] > 2:
                    val = update_cell_value(row[col_dict_df["CELL"]])
                    main_output[site].append(val)
        if len(main_output[site]) > 0:
            main_output[site] = ["High SDCCH Blocking at %s" % str(", ".join(main_output[site]))]
        final_output[site].append(main_output[site])
        
    main_output = {}
    for site in site_list:
        main_output[site] = []
        for row in fdataset.iloc:
            if row[col_dict_df["SITE"]] == site:
                if row[col_dict_df["TCH Traffic"]] == 0:
                    val = update_cell_value(row[col_dict_df["CELL"]])
                    main_output[site].append(val)
        if len(main_output[site]) > 0:
            main_output[site] = ["0 TCH Traffic at %s" % str(", ".join(main_output[site]))]
        final_output[site].append(main_output[site])
    
    main_output = {}
    for site in site_list:
        main_output[site] = []
        for row in fdataset.iloc:
            if row[col_dict_df["SITE"]] == site:
                if row[col_dict_df["SDCCH Traffic"]] == 0:
                    val = update_cell_value(row[col_dict_df["CELL"]])
                    main_output[site].append(val)
        if len(main_output[site]) > 0:
            main_output[site] = ["0 SDCCH Traffic at %s" % str(", ".join(main_output[site]))]
        final_output[site].append(main_output[site])
    
    main_output = {}
    for site in site_list:
        main_output[site] = []
        for row in fdataset.iloc:
            if row[col_dict_df["SITE"]] == site:
                if row[col_dict_df["TCH Drop Call Rate"]] > 2 and row[col_dict_df["TCH Drop Calls"]] > 10:
                    val = update_cell_value(row[col_dict_df["CELL"]])
                    main_output[site].append(val)
        if len(main_output[site]) > 0:
            main_output[site] = ["High TCH Drop Call Rate at %s" % str(", ".join(main_output[site]))]
        final_output[site].append(main_output[site])
        
    main_output = {}
    for site in site_list:
        main_output[site] = []
        for row in fdataset.iloc:
            if row[col_dict_df["SITE"]] == site:
                if row[col_dict_df["SDCCH Drop Ratio"]] > 2:
                    val = update_cell_value(row[col_dict_df["CELL"]])
                    main_output[site].append(val)
        if len(main_output[site]) > 0:
            main_output[site] = ["High SDCCH Drop Ratio at %s" % str(", ".join(main_output[site]))]
        final_output[site].append(main_output[site])
        
    for key, value in final_output.items():
        output = []
        for line in value:
            output += line
        final_output[key] = ["\n".join(output)]
        
    for k,v in final_output.items():
        final_output[k]=formatpara(''.join(final_output[k]))
    
    return final_output


def generate_umts_remarks(site_list,fdataset,col_dict_df):
    final_output = {}
        
    main_output = {}
    for site in site_list:
        main_output[site] = []
        final_output[site] = []
        for row in fdataset.iloc:
            if row[col_dict_df["SITE"]] == site:
                if row[col_dict_df["Average RTWP"]] > -107 and row[col_dict_df["Average RTWP"]] < -102:
                    val = update_cell_value(row[col_dict_df["CELL"]])
                    main_output[site].append(val)
        if len(main_output[site]) > 0:
            main_output[site] = ["Average RTWP lies outside the recommended threshold at %s" % str(", ".join(main_output[site]))]
        final_output[site].append(main_output[site])
    
    main_output = {}
    for site in site_list:
        main_output[site] = []
        final_output[site] = []
        for row in fdataset.iloc:
            if row[col_dict_df["SITE"]] == site:
                if row[col_dict_df["Cell Availability Excluding Blk By User"]] == 0:
                    val = update_cell_value(row[col_dict_df["CELL"]])
                    main_output[site].append(val)
        if len(main_output[site]) > 0:
            main_output[site] = ["0 Cell Availability at %s" % str(", ".join(main_output[site]))]
        final_output[site].append(main_output[site]) 
        
    main_output = {}
    for site in site_list:
        main_output[site] = []
        for row in fdataset.iloc:
            if row[col_dict_df["SITE"]] == site:
                if row[col_dict_df["Cell Availability Excluding Blk By User"]] > 0 and row[col_dict_df["Cell Availability Excluding Blk By User"]] < 98:
                    val = update_cell_value(row[col_dict_df["CELL"]])
                    main_output[site].append(val)
        if len(main_output[site]) > 0:
            main_output[site] = ["Low Cell Availability at %s" % str(", ".join(main_output[site]))]
        final_output[site].append(main_output[site])
        
    main_output = {}
    for site in site_list:
        main_output[site] = []
        for row in fdataset.iloc:
            if row[col_dict_df["SITE"]] == site:
                if row[col_dict_df["HSDPA Resource Drop rate for NRT Traffic (Int & Bgr)"]] > 2:
                    val = update_cell_value(row[col_dict_df["CELL"]])
                    main_output[site].append(val)
        if len(main_output[site]) > 0:
            main_output[site] = ["High HSDPA Resource Drop rate at %s" % str(", ".join(main_output[site]))]
        final_output[site].append(main_output[site])
        
    main_output = {}
    for site in site_list:
        main_output[site] = []
        for row in fdataset.iloc:
            if row[col_dict_df["SITE"]] == site:
                if row[col_dict_df["PS RAB Drop Rate (GCR)"]] > 2:
                    val = update_cell_value(row[col_dict_df["CELL"]])
                    main_output[site].append(val)
        if len(main_output[site]) > 0:
            main_output[site] = ["High PS RAB Drop Rate at %s" % str(", ".join(main_output[site]))]
        final_output[site].append(main_output[site])
    
    main_output = {}
    for site in site_list:
        main_output[site] = []
        for row in fdataset.iloc:
            if row[col_dict_df["SITE"]] == site:
                if row[col_dict_df["Voice Drop Rate (GCR)"]] > 2:
                    val = update_cell_value(row[col_dict_df["CELL"]])
                    main_output[site].append(val)
        if len(main_output[site]) > 0:
            main_output[site] = ["High Voice Drop Rate at %s" % str(", ".join(main_output[site]))]
        final_output[site].append(main_output[site])
        
    main_output = {}
    for site in site_list:
        main_output[site] = []
        for row in fdataset.iloc:
            if row[col_dict_df["SITE"]] == site:
                if row[col_dict_df["Voice Traffic Erlangs"]] == 0:
                    val = update_cell_value(row[col_dict_df["CELL"]])
                    main_output[site].append(val)
        if len(main_output[site]) > 0:
            main_output[site] = ["0 Voice Traffic at %s" % str(", ".join(main_output[site]))]
        final_output[site].append(main_output[site])
        
        
    for key, value in final_output.items():
        output = []
        for line in value:
            output += line
        final_output[key] = ["\n".join(output)]
        
    for k,v in final_output.items():
        final_output[k]=formatpara(''.join(final_output[k]))
    
    return final_output


import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill,Alignment
from itertools import groupby
# from glob import glob
# from openpyxl.cell import Cell

# file_name_rem=input("Enter the remarks file name: ")
file_name_rem=''
for f_name in os.listdir(os.getcwd()):
    if f_name.endswith('Remarks.xlsx'):
        file_name_rem=f_name
        print(file_name_rem)
        break
    
remarks_df=pd.read_excel(file_name_rem)
cross_check = ["L1900_1C", "L1900_2C", "L2500_1C", "L2500_2C"]

site_totaltechs_list={}
for file_name in os.listdir(os.getcwd()):
    if file_name.endswith('hr.xlsx'):
        df = pd.read_excel(file_name)
        site_totaltechs_list = create_all_antennas(df, site_totaltechs_list)  
    if file_name.endswith('hr.csv'):
        df = pd.read_csv(file_name)
        site_totaltechs_list = create_all_antennas(df, site_totaltechs_list) 

manual_dict={}
for key in site_totaltechs_list:

    site_info = remarks_df.loc[remarks_df['Site ID'] == key, 'New Tech']

    if len(site_info):
        try:
            new_tech = [tech.strip() for tech in site_info.to_list()[0].split(',')] # New tech
            total = site_totaltechs_list[key]
            leftout = [tech for tech in cross_check if tech not in (total+new_tech)]
            if leftout:
                #print('Need to check manually for {} : {}'.format(key, leftout))
                leftout=','.join(leftout)
                manual_dict[key]=leftout

            locked_tech = [tech for tech in new_tech if tech not in total]
            #locked_tech = [tech for tech in locked_tech if not tech.startswith('5G')]
            remarks_df.loc[remarks_df['Site ID'] == key, 'Locked Tech'] = ', '.join(list(locked_tech))

        except Exception as e:
            pass

lock_techs = {x[0] : x[1] for x in remarks_df.loc[:, ['Site ID', 'Locked Tech']].fillna('').to_dict('split')['data']}


wb=load_workbook(file_name_rem)
ws=wb.active
#ws=wb["Sheet1"]
# ws2=wb["Daily"]

neongreenFill=PatternFill(start_color='39ff14', #for pass sites
                   end_color='39ff14',
                   fill_type='solid')
goldFill=PatternFill(start_color='ffd700', #for pending sites
                   end_color='ffd700',
                   fill_type='solid')
redFill=PatternFill(start_color='ff3333', #for fail sites
                   end_color='ff3333',
                   fill_type='solid')
orangeFill=PatternFill(start_color='ff8c00', #for overdue sites
                   end_color='ff8c00',
                   fill_type='solid')

dim1=ws.dimensions
#ws.insert_cols(ord(dim1[3])+1-64+1)
nextcol8=chr(ord(dim1[3])+1+1+1+1+1+1+1+1) #for locked techs
ws.insert_cols(ord(dim1[3])+1-64+1+1+1+1+1+1+1+1) #for locked techs
ws[nextcol8+'1']='Locked Tech'
# dim1=''
# dim1=ws.dimensions
# col_dict_rem={}
# j=0
# for i in range(65,ord(dim1[3])+1): #remarks file column names dictionary
#     col_dict_rem[ws[chr(i)+'1'].value]=j
#     j+=1

# ws.insert_cols(col_dict_rem['New Tech']+2) #inserting a column after 'New Tech'
# ws[str(chr(col_dict_rem['New Tech']+65+1)+str('1'))]='Locked Tech' #this new column is 'Locked Tech'

# dim1=''
# dim1=ws.dimensions
# col_dict_rem={} #rerunning because new column is now added
# j=0
# for i in range(65,ord(dim1[3])+1): #remarks file column names dictionary
#     col_dict_rem[ws[chr(i)+'1'].value]=j
#     j+=1

# dim1=''
# dim1=ws.dimensions #rerunning because new column is now added
nextcol1=chr(ord(dim1[3])+1) #for LTE new tech remarks
nextcol2=chr(ord(dim1[3])+1+1) #for 5G new tech remarks
nextcol3=chr(ord(dim1[3])+1+1+1) #for GSM new tech remarks
nextcol4=chr(ord(dim1[3])+1+1+1+1) #for UMTS new tech remarks
nextcol5=chr(ord(dim1[3])+1+1+1+1+1) #for existing tech remarks
nextcol6=chr(ord(dim1[3])+1+1+1+1+1+1) #for total tech remarks
nextcol7=chr(ord(dim1[3])+1+1+1+1+1+1+1) #Blank column
#nextcol5=chr(ord(dim1[3])+1+1+1+1+1+1) #for locked techs
nextcol9=chr(ord(dim1[3])+1+1+1+1+1+1+1+1+1) #for need to check manual sites
trows1=int(dim1[4::])
# ws.insert_cols(ord(dim1[3])+1-64+1) #for LTE new tech remarks
# ws.insert_cols(ord(dim1[3])+1-64+1+1) #for 5G new tech remarks
# ws.insert_cols(ord(dim1[3])+1-64+1+1+1) #for existing tech remarks
# ws.insert_cols(ord(dim1[3])+1-64+1+1+1+1) #for total tech remarks
# ws.insert_cols(ord(dim1[3])+1-64+1+1+1+1+1) #Blank column
#ws.insert_cols(ord(dim1[3])+1-64+1+1+1+1+1+1) #for locked techs
ws.insert_cols(ord(dim1[3])+1-64+1+1+1+1+1+1+1+1+1) #for need to check manual sites
ws[nextcol1+'1']='New Tech LTE KPI Validation Remarks'
ws[nextcol2+'1']='New Tech 5G KPI Validation Remarks'
ws[nextcol3+'1']='New Tech GSM KPI Validation Remarks'
ws[nextcol4+'1']='New Tech UMTS KPI Validation Remarks'
ws[nextcol5+'1']='Existing Tech KPI Validation Remarks'
ws[nextcol6+'1']='Total Tech KPI Validation Remarks'
ws[nextcol7+'1']=''
#ws[nextcol5+'1']='Locked Tech'
ws[nextcol9+'1']='Need to check manually for the below techs'
for i in range (2,trows1+1):
    ws[nextcol1+str(i)].value='Site ID not found'
    ws[nextcol2+str(i)].value='Site ID not found'
    ws[nextcol3+str(i)].value='Site ID not found'
    ws[nextcol4+str(i)].value='Site ID not found'
    ws[nextcol5+str(i)].value='Site ID not found'
    ws[nextcol6+str(i)].value='Site ID not found'

dim2='' #ONLY for the purpose of calculating col_dict_rem
dim2=ws.dimensions
col_dict_rem={}
j=0
for i in range(65,ord(dim2[3])+1): #remarks file column names dictionary
    col_dict_rem[ws[chr(i)+'1'].value]=j
    j+=1


ntc=chr(65+int(col_dict_rem['New Tech'])) #newtechcolumn
ltc=chr(65+int(col_dict_rem['Locked Tech'])) #lockedtech column
mtc=chr(65+int(col_dict_rem['Need to check manually for the below techs'])) #manuallockedtechcheck column
site_techs_list={}
ns_string='' #site techs after new-locked to be converted into string
for i in range (2,trows1+1):
    ns=str(ws[ntc+str(i)].value).split(',')
    for j in range(0,len(ns)):
        ns[j]=ns[j].strip()
    # for j in ns: #to remove GSM and UMTS techs from further processing in the system as we aren't doing KPI monitoring of these techs as of now
    #    if j=='GSM' or j=='U1900' or j=='U2100':
    #        ns.remove(j)
    ls=str(ws[ltc+str(i)].value).split(',')
    for j in range(0,len(ls)):
        ls[j]=ls[j].strip()
    for a in ls:
        for b in ns:
            if a==b:
                ns.remove(a)
    ns_string=', '.join(ns)
    site_techs_list[ws['A'+str(i)].value]=ns_string
    ns_string=''

# Fill locked tech column from lock_techs
for kl,vl in lock_techs.items():
    for i in range(2,trows1+1):
        if kl==ws[str(chr(col_dict_rem['Site ID']+65))+str(i)].value:
            if vl=='': #or vl=='GSM' or vl=='U1900' or vl=='U2100'
                continue
            else:
                ws[ltc+str(i)]=vl
# Fill Need to check manually for the below techs column from manual_dict
for km,vm in manual_dict.items():
    for i in range(2,trows1+1):
        if km==ws[str(chr(col_dict_rem['Site ID']+65))+str(i)].value:
            if vm=='':
                continue
            else:
                ws[mtc+str(i)]=str(vm)


# for file_name in glob(fr'{file_path}\*.xlsx'):
for file_name in os.listdir(os.getcwd()):
    if file_name.endswith('hr.xlsx') or file_name.endswith('hr.csv'):
        print('Reading files...')
        file_duration=''
        for i in file_name: #to determine which file is for what duration KPI
            if (i.isdigit()==1):
                file_duration+=i
        file_duration=int(file_duration)
        
        dataset = pd.DataFrame()
        if file_name.endswith('hr.xlsx'):
            dataset = pd.read_excel(file_name)  #excel file import
        if file_name.endswith('hr.csv'):
            dataset = pd.read_csv(file_name)  #csv file import
        dataset.sort_values(by=['SITE','CELL'], inplace=True, ignore_index=True)
        # dataset = pd.read_csv(file_name)  #csv file import
        col_dict_df={}
        ind=0
        for i in list(dataset.columns): #to determine site file column names dictionary
            col_dict_df[i]=ind
            ind+=1
        site_list = list(set(list(dataset.iloc[:, col_dict_df['SITE']].values)))
        
        #new tech filter
        
        dataset2=dataset
        for i in range (0,len(dataset2)):
            dataset2=dataset2.drop(i,axis=0)
        dataset3=dataset
        for i in range (0,len(dataset3)):
            dataset3=dataset3.drop(i,axis=0)
        
        for site in site_list:
            nt=site_techs_list[site].split(', ') #list of new techs-locked techs for that particular site
            for s in nt:
                if s=='NR600' or s=='NR2500':
                    newtech=s[0:6]
                else:
                    newtech=s[0:6]
                newtech=newtech.strip()
                newtech=newtech.strip('_')
                if s.find('_1C',0,len(s))!=-1:
                    c1=s[-3::]
                    for k,v in carrier_names.items():
                        if v[:]==c1:
                            c1=k
                if s.find('_2C',0,len(s))!=-1:
                    c2=s[-3::]
                    for k,v in carrier_names.items():
                        if v[:]==c2:
                            c2=k
                if s.find('_3C',0,len(s))!=-1:
                    c3=s[-3::]
                    for k,v in carrier_names.items():
                        if v[:]==c3:
                            c3=k
                for i in range (0,len(dataset)):
                    if dataset['SITE'][i]==site:
                        for ki,vi in tech_names.items():   
                            if vi==newtech:
                                if s.find('_1C',0,len(s))!=-1:
                                    if (dataset['CELL'][i][0]==ki and dataset['CELL'][i][-1]==c1):
                                        dataset2=dataset2.append(dataset.iloc[i,:])
                                elif s.find('_2C',0,len(s))!=-1:
                                    if (dataset['CELL'][i][0]==ki and dataset['CELL'][i][-1]==c2):
                                        dataset2=dataset2.append(dataset.iloc[i,:])
                                elif s.find('_3C',0,len(s))!=-1:
                                    if (dataset['CELL'][i][0]==ki and dataset['CELL'][i][-1]==c3):
                                        dataset2=dataset2.append(dataset.iloc[i,:])
                                else:
                                    if dataset['CELL'][i][0]==ki:
                                        dataset2=dataset2.append(dataset.iloc[i,:])
     
        #Now dataset2 would be referred to generate remarks for new techs
        #Now dataset3 would be referred to generate remarks for existing techs
        
        dataset3=pd.concat([dataset,dataset2]).drop_duplicates(keep=False)
        #Now dataset would be referred to generate remarks for all techs available in the site file
        
        if (file_name.find('LTE')!=-1):
        
            final_output=generate_lte_remarks(site_list,dataset2,col_dict_df)
                
            #print(final_output)
            df = pd.DataFrame.from_dict(final_output, orient="Index")
            k=''
            for i in list(df.index):
                for j in range(2,trows1+1):
                    if i == ws['A'+str(j)].value:
                        if (str((df.loc[i,0]))!='') and (ws[nextcol1+str(j)].value=='Site ID not found') or (ws[nextcol1+str(j)].value=='') or (ws[nextcol1+str(j)].value=='None'): # or (ws[nextcol1+str(j)].value.find('issue was observed')!=-1):
                            ws[nextcol1+str(j)] = 'During '+str(ws[str(chr((col_dict_rem['Hour'])+65))+str(j)].value)+' hour KPI validation, the following issues were observed:'
                            ws[nextcol1+str(j)] = ws[nextcol1+str(j)].value+'\n'+str((df.loc[i,0]))
                        if str((df.loc[i,0]))=='':
                            for a,b in duration_meaning.items():
                                if int(b)==file_duration:
                                    k=a
                                    for c,d in col_dict_rem.items():
                                        if k==c:
                                            if not ws[chr(col_dict_rem['Locked Tech']+65)+str(j)].value: #check if blank or None
                                                # ws[str(chr(65+int(d)))+str(j)]='Pass'
                                                # ws[str(chr(65+int(d)))+str(j)].fill=neongreenFill
                                                if (ws[nextcol1+str(j)].value=='Site ID not found') or (ws[nextcol1+str(j)].value=='') or (ws[nextcol1+str(j)].value=='None'):                                                    
                                                    ws[nextcol1+str(j)]='During '+str(ws[str(chr((col_dict_rem['Hour'])+65))+str(j)].value)+' hour KPI validation, no LTE issue was observed'
                                                    if not ws[str(chr(65+int(d)))+str(j)].value or ws[str(chr(65+int(d)))+str(j)].fill.start_color.index=='00000000':
                                                        ws[str(chr(65+int(d)))+str(j)]='Pass'
                                                        ws[str(chr(65+int(d)))+str(j)].fill=neongreenFill
                                                # elif (ws[nextcol1+str(j)].value.find('issue was observed')!=-1):
                                                #         ws[str(chr(65+int(d)))+str(j)]='Pass'
                                                #         ws[str(chr(65+int(d)))+str(j)].fill=neongreenFill
                                                #         ws[nextcol1+str(j)]=ws[nextcol1+str(j)].value+'During '+str(ws[str(chr((col_dict_rem['Hour'])+65))+str(j)].value)+' hour KPI validation, no LTE issue was observed'
                                                # else:
                                                #     ws[str(chr(65+int(d)))+str(j)]='Fail'
                                                #     ws[str(chr(65+int(d)))+str(j)].fill=redFill
                                                #     ws[nextcol1+str(j)]=ws[nextcol1+str(j)].value+'\n'+'During '+str(ws[str(chr((col_dict_rem['Hour'])+65))+str(j)].value)+' hour KPI validation, no LTE issue was observed'
                                            else:
                                                if (ws[nextcol1+str(j)].value=='Site ID not found') or (ws[nextcol1+str(j)].value=='') or (ws[nextcol1+str(j)].value=='None'):
                                                    ws[nextcol1+str(j)]='During '+str(ws[str(chr((col_dict_rem['Hour'])+65))+str(j)].value)+' hour KPI validation, no LTE issue was observed'
                                                else:
                                                    ws[nextcol1+str(j)]=ws[nextcol1+str(j)].value+'\n'+'During '+str(ws[str(chr((col_dict_rem['Hour'])+65))+str(j)].value)+' hour KPI validation, no LTE issue was observed'
                                                ws[str(chr(65+int(d)))+str(j)]='Overdue'
                                                ws[str(chr(65+int(d)))+str(j)].fill=orangeFill 
                                                # ll=str(ws[str(chr((col_dict_rem['Locked Tech'])+65))+str(j)].value).split(', ') #locked list
                                                # otj='' #overdue tech join
                                                # for il in ll:
                                                #     if il!='NR2500' and il!='NR600' and il!='GSM' and il!='U1900' and il!='U2100':
                                                #         otj=', '.join([otj,il])
                                                # otj=otj.strip(', ')
                                                # if not otj:
                                                #     break
                                                # else:
                                                #     if (ws[nextcol1+str(j)].value=='Site ID not found') or (ws[nextcol1+str(j)].value=='') or (ws[nextcol1+str(j)].value=='None'):
                                                #         ws[nextcol1+str(j)]='KPI validation is pending as New Tech '+otj+' is locked'
                                                #     else:
                                                #         ws[nextcol1+str(j)]=ws[nextcol1+str(j)].value+'\n'+'KPI validation is pending as New Tech '+otj+' is locked'
                                            for pe in range(int(d)+1,col_dict_rem['KPI Next 2 Days Validated']+1):
                                                ws[str(chr(65+int(pe)))+str(j)]='Pending'
                                                ws[str(chr(65+int(pe)))+str(j)].fill=goldFill
                        # ll=str(ws[str(chr((col_dict_rem['Locked Tech'])+65))+str(j)].value).split(', ') #locked list
                        # otj='' #overdue tech join
                        # for il in ll:
                        #     if il!='NR2500' and il!='NR600' and il!='GSM' and il!='U1900' and il!='U2100':
                        #         otj=', '.join([otj,il])
                        #         otj=otj.strip(', ')
                        #         if not otj:
                        #             break
                        #         else:
                        #             if (ws[nextcol1+str(j)].value=='Site ID not found') or (ws[nextcol1+str(j)].value=='') or (ws[nextcol1+str(j)].value=='None'):
                        #                 ws[nextcol1+str(j)]='KPI validation is pending as New Tech '+otj+' is locked'
                        #             else:
                        #                 ws[nextcol1+str(j)]=ws[nextcol1+str(j)].value+'\n'+'KPI validation is pending as New Tech '+otj+' is locked'
                                                
                                                   
                        elif str((df.loc[i,0]))!='':
                            for a,b in duration_meaning.items():
                                if int(b)==file_duration:
                                    k=a
                                    for c,d in col_dict_rem.items():
                                        if k==c:
                                            if not ws[chr(col_dict_rem['Locked Tech']+65)+str(j)].value:
                                                ws[str(chr(65+int(d)))+str(j)]='Fail'
                                                ws[str(chr(65+int(d)))+str(j)].fill=redFill
                                            else:
                                                ws[str(chr(65+int(d)))+str(j)]='Overdue'
                                                ws[str(chr(65+int(d)))+str(j)].fill=orangeFill
                                            for pe in range(int(d)+1,col_dict_rem['KPI Next 2 Days Validated']+1):
                                                ws[str(chr(65+int(pe)))+str(j)]='Pending'
                                                ws[str(chr(65+int(pe)))+str(j)].fill=goldFill
                                                
                        ll=str(ws[str(chr((col_dict_rem['Locked Tech'])+65))+str(j)].value).split(', ') #locked list
                        otj='' #overdue tech join
                        for il in ll:
                            fp=il.split('_')[0] #fp is first part
                            if fp!='NR2500' and fp!='NR600' and fp!='GSM' and fp!='U1900' and fp!='U2100':
                                otj=', '.join([otj,il])
                        otj=otj.strip(', ')
                        if otj=='None' or otj==None or otj=='':
                            break
                        else:
                            if (ws[nextcol1+str(j)].value=='Site ID not found') or (ws[nextcol1+str(j)].value=='') or (ws[nextcol1+str(j)].value=='None'):
                                ws[nextcol1+str(j)]='KPI validation is pending as New Tech '+otj+' is locked'
                            else:
                                ws[nextcol1+str(j)]=ws[nextcol1+str(j)].value+'\n'+'KPI validation is pending as New Tech '+otj+' is locked'
            
            
            final_output=generate_lte_remarks(site_list,dataset3,col_dict_df)
                
            #print(final_output)
            df = pd.DataFrame.from_dict(final_output, orient="Index")
            k=''
            for i in list(df.index):
                for j in range(2,trows1+1):
                    if i == ws['A'+str(j)].value:
                        if (ws[nextcol5+str(j)].value=='Site ID not found') or (ws[nextcol5+str(j)].value=='') or (ws[nextcol5+str(j)].value=='None'): # or (ws[nextcol5+str(j)].value.find('issue was observed')!=-1):
                            ws[nextcol5+str(j)] = str((df.loc[i,0]))
                        else:
                            ws[nextcol5+str(j)] = ws[nextcol5+str(j)].value+'\n'+str((df.loc[i,0]))
                        if str((df.loc[i,0]))=='':
                            for a,b in duration_meaning.items():
                                if int(b)==file_duration:
                                    k=a
                                    for c,d in col_dict_rem.items():
                                        if k==c:
                                            if not ws[chr(col_dict_rem['Locked Tech']+65)+str(j)].value: #check if blank or None
                                            # ws[str(chr(65+int(d)))+str(j)]='Pass'
                                            # ws[str(chr(65+int(d)))+str(j)].fill=neongreenFill
                                                if (ws[nextcol5+str(j)].value=='Site ID not found') or (ws[nextcol5+str(j)].value=='') or (ws[nextcol5+str(j)].value=='None'):
                                                    ws[nextcol5+str(j)]='During '+str(ws[str(chr((col_dict_rem['Hour'])+65))+str(j)].value)+' hour KPI validation, no LTE issue was observed'
                                                else:
                                                    ws[nextcol5+str(j)]=ws[nextcol5+str(j)].value+'\n'+'During '+str(ws[str(chr((col_dict_rem['Hour'])+65))+str(j)].value)+' hour KPI validation, no LTE issue was observed'
                                           # else:
                                               # ws[str(chr(65+int(d)))+str(j)]='Overdue'
                                               # ws[str(chr(65+int(d)))+str(j)].fill=orangeFill
                                               # ws[nextcol5+str(j)]='KPI validation is pending as New Tech '+str(ws[str(chr((col_dict_rem['Locked Tech'])+65))+str(j)].value)+' is locked/down'
                                           # for pe in range(int(d)+1,col_dict_rem['KPI Next 2 Days Validated']+1):
                                           #     ws[str(chr(65+int(pe)))+str(j)]='Pending'
                                           #     ws[str(chr(65+int(pe)))+str(j)].fill=goldFill
                        # elif str((df.loc[i,0]))!='':
                        #     for a,b in duration_meaning.items():
                        #         if int(b)==file_duration:
                        #             k=a
                        #             for c,d in col_dict_rem.items():
                        #                if k==c:
                        #                    if not ws[chr(col_dict_rem['Locked Tech']+65)+str(j)].value:
                        #                        ws[str(chr(65+int(d)))+str(j)]='Fail'
                        #                        ws[str(chr(65+int(d)))+str(j)].fill=redFill
                        #                    else:
                        #                        ws[str(chr(65+int(d)))+str(j)]='Overdue'
                        #                        ws[str(chr(65+int(d)))+str(j)].fill=orangeFill
                        #                    for pe in range(int(d)+1,col_dict_rem['KPI Next 2 Days Validated']+1):
                        #                        ws[str(chr(65+int(pe)))+str(j)]='Pending'
                        #                        ws[str(chr(65+int(pe)))+str(j)].fill=goldFill
                        # break
                    
                # for k in range(2,trows2+1):
                #     if i == ws2['A'+str(k)].value:
                #         ws2[nextcol5+str(k)] = str((df.loc[i,0]))
                
            final_output=generate_lte_remarks(site_list,dataset,col_dict_df)
                
            #print(final_output)
            df = pd.DataFrame.from_dict(final_output, orient="Index")
            k=''
            for i in list(df.index):
                for j in range(2,trows1+1):
                    if i == ws['A'+str(j)].value:
                        if (ws[nextcol6+str(j)].value=='Site ID not found') or (ws[nextcol6+str(j)].value=='') or (ws[nextcol6+str(j)].value=='None'): # or (ws[nextcol6+str(j)].value.find('issue was observed')!=-1):
                            ws[nextcol6+str(j)] = str((df.loc[i,0]))
                        else:
                            ws[nextcol6+str(j)] = ws[nextcol6+str(j)].value+'\n'+str((df.loc[i,0]))
                        if str((df.loc[i,0]))=='':
                            for a,b in duration_meaning.items():
                                if int(b)==file_duration:
                                    k=a
                                    for c,d in col_dict_rem.items():
                                        if k==c:
                                            if not ws[chr(col_dict_rem['Locked Tech']+65)+str(j)].value: #check if blank or None
                                                # ws[str(chr(65+int(d)))+str(j)]='Pass'
                                                # ws[str(chr(65+int(d)))+str(j)].fill=neongreenFill
                                                if (ws[nextcol6+str(j)].value=='Site ID not found') or (ws[nextcol6+str(j)].value=='') or (ws[nextcol6+str(j)].value=='None'):
                                                    ws[nextcol6+str(j)]='During '+str(ws[str(chr((col_dict_rem['Hour'])+65))+str(j)].value)+' hour KPI validation, no LTE issue was observed'
                                                else:
                                                    ws[nextcol6+str(j)]=ws[nextcol6+str(j)].value+'\n'+'During '+str(ws[str(chr((col_dict_rem['Hour'])+65))+str(j)].value)+' hour KPI validation, no LTE issue was observed'
                                               # else:
                                                   # ws[str(chr(65+int(d)))+str(j)]='Overdue'
                                                   # ws[str(chr(65+int(d)))+str(j)].fill=orangeFill
                                                   # ws[nextcol6+str(j)]='KPI validation is pending as New Tech '+str(ws[str(chr((col_dict_rem['Locked Tech'])+65))+str(j)].value)+' is locked/down'
                                               # for pe in range(int(d)+1,col_dict_rem['KPI Next 2 Days Validated']+1):
                                               #     ws[str(chr(65+int(pe)))+str(j)]='Pending'
                                               #     ws[str(chr(65+int(pe)))+str(j)].fill=goldFill
                            # elif str((df.loc[i,0]))!='':
                            #     for a,b in duration_meaning.items():
                            #         if int(b)==file_duration:
                            #             k=a
                            #             for c,d in col_dict_rem.items():
                            #                if k==c:
                            #                    if not ws[chr(col_dict_rem['Locked Tech']+65)+str(j)].value:
                            #                        ws[str(chr(65+int(d)))+str(j)]='Fail'
                            #                        ws[str(chr(65+int(d)))+str(j)].fill=redFill
                            #                    else:
                            #                        ws[str(chr(65+int(d)))+str(j)]='Overdue'
                            #                        ws[str(chr(65+int(d)))+str(j)].fill=orangeFill
                            #                    for pe in range(int(d)+1,col_dict_rem['KPI Next 2 Days Validated']+1):
                            #                        ws[str(chr(65+int(pe)))+str(j)]='Pending'
                            #                        ws[str(chr(65+int(pe)))+str(j)].fill=goldFill
                            # break
                    
        
        
        if (file_name.find('FIVEG')!=-1):
        
            final_output=generate_5g_remarks(site_list,dataset2,col_dict_df)
                
            #print(final_output)
            df = pd.DataFrame.from_dict(final_output, orient="Index")
            k=''
            for i in list(df.index):
                for j in range(2,trows1+1):
                    if i == ws['A'+str(j)].value:
                        if (str((df.loc[i,0]))!='') and (ws[nextcol2+str(j)].value=='Site ID not found') or (ws[nextcol2+str(j)].value=='') or (ws[nextcol2+str(j)].value=='None'): # or (ws[nextcol2+str(j)].value.find('issue was observed')!=-1):
                            ws[nextcol2+str(j)] = 'During '+str(ws[str(chr((col_dict_rem['Hour'])+65))+str(j)].value)+' hour KPI validation, the following issues were observed:'
                            ws[nextcol2+str(j)] = ws[nextcol2+str(j)].value+'\n'+str((df.loc[i,0]))
                        if str((df.loc[i,0]))=='':
                            for a,b in duration_meaning.items():
                                if int(b)==file_duration:
                                    k=a
                                    for c,d in col_dict_rem.items():
                                        if k==c:
                                            if not ws[chr(col_dict_rem['Locked Tech']+65)+str(j)].value: #check if blank or None
                                                # ws[str(chr(65+int(d)))+str(j)]='Pass'
                                                # ws[str(chr(65+int(d)))+str(j)].fill=neongreenFill
                                                if (ws[nextcol2+str(j)].value=='Site ID not found') or (ws[nextcol2+str(j)].value=='') or (ws[nextcol2+str(j)].value=='None'):
                                                    ws[nextcol2+str(j)]='During '+str(ws[str(chr((col_dict_rem['Hour'])+65))+str(j)].value)+' hour KPI validation, no 5G issue was observed'
                                                    if not ws[str(chr(65+int(d)))+str(j)].value or ws[str(chr(65+int(d)))+str(j)].fill.start_color.index=='00000000':
                                                        ws[str(chr(65+int(d)))+str(j)]='Pass'
                                                        ws[str(chr(65+int(d)))+str(j)].fill=neongreenFill
                                                # elif (ws[nextcol2+str(j)].value.find('issue was observed')!=-1):
                                                #         ws[str(chr(65+int(d)))+str(j)]='Pass'
                                                #         ws[str(chr(65+int(d)))+str(j)].fill=neongreenFill
                                                #         ws[nextcol2+str(j)]=ws[nextcol2+str(j)].value+'During '+str(ws[str(chr((col_dict_rem['Hour'])+65))+str(j)].value)+' hour KPI validation, no 5G issue was observed'
                                                # else:
                                                #     ws[str(chr(65+int(d)))+str(j)]='Fail'
                                                #     ws[str(chr(65+int(d)))+str(j)].fill=redFill
                                                #     ws[nextcol2+str(j)]=ws[nextcol2+str(j)].value+'\n'+'During '+str(ws[str(chr((col_dict_rem['Hour'])+65))+str(j)].value)+' hour KPI validation, no 5G issue was observed'
                                            else:
                                                if (ws[nextcol2+str(j)].value=='Site ID not found') or (ws[nextcol2+str(j)].value=='') or (ws[nextcol2+str(j)].value=='None'):
                                                    ws[nextcol2+str(j)]='During '+str(ws[str(chr((col_dict_rem['Hour'])+65))+str(j)].value)+' hour KPI validation, no 5G issue was observed'
                                                else:
                                                    ws[nextcol2+str(j)]=ws[nextcol2+str(j)].value+'\n'+'During '+str(ws[str(chr((col_dict_rem['Hour'])+65))+str(j)].value)+' hour KPI validation, no 5G issue was observed'
                                                ws[str(chr(65+int(d)))+str(j)]='Overdue'
                                                ws[str(chr(65+int(d)))+str(j)].fill=orangeFill 
                                                # ll=str(ws[str(chr((col_dict_rem['Locked Tech'])+65))+str(j)].value).split(', ') #locked list
                                                # otj='' #overdue tech join
                                                # for il in ll:
                                                #     if il!='NR2500' and il!='NR600' and il!='GSM' and il!='U1900' and il!='U2100':
                                                #         otj=', '.join([otj,il])
                                                # otj=otj.strip(', ')
                                                # if not otj:
                                                #     break
                                                # else:
                                                #     if (ws[nextcol2+str(j)].value=='Site ID not found') or (ws[nextcol2+str(j)].value=='') or (ws[nextcol2+str(j)].value=='None'):
                                                #         ws[nextcol2+str(j)]='KPI validation is pending as New Tech '+otj+' is locked'
                                                #     else:
                                                #         ws[nextcol2+str(j)]=ws[nextcol2+str(j)].value+'\n'+'KPI validation is pending as New Tech '+otj+' is locked'
                                            for pe in range(int(d)+1,col_dict_rem['KPI Next 2 Days Validated']+1):
                                                ws[str(chr(65+int(pe)))+str(j)]='Pending'
                                                ws[str(chr(65+int(pe)))+str(j)].fill=goldFill
                        # ll=str(ws[str(chr((col_dict_rem['Locked Tech'])+65))+str(j)].value).split(', ') #locked list
                        # otj='' #overdue tech join
                        # for il in ll:
                        #     if il!='NR2500' and il!='NR600' and il!='GSM' and il!='U1900' and il!='U2100':
                        #         otj=', '.join([otj,il])
                        #         otj=otj.strip(', ')
                        #         if not otj:
                        #             break
                        #         else:
                        #             if (ws[nextcol2+str(j)].value=='Site ID not found') or (ws[nextcol2+str(j)].value=='') or (ws[nextcol2+str(j)].value=='None'):
                        #                 ws[nextcol2+str(j)]='KPI validation is pending as New Tech '+otj+' is locked'
                        #             else:
                        #                 ws[nextcol2+str(j)]=ws[nextcol2+str(j)].value+'\n'+'KPI validation is pending as New Tech '+otj+' is locked'
                                                
                                                   
                        elif str((df.loc[i,0]))!='':
                            for a,b in duration_meaning.items():
                                if int(b)==file_duration:
                                    k=a
                                    for c,d in col_dict_rem.items():
                                        if k==c:
                                            if not ws[chr(col_dict_rem['Locked Tech']+65)+str(j)].value:
                                                ws[str(chr(65+int(d)))+str(j)]='Fail'
                                                ws[str(chr(65+int(d)))+str(j)].fill=redFill
                                            else:
                                                ws[str(chr(65+int(d)))+str(j)]='Overdue'
                                                ws[str(chr(65+int(d)))+str(j)].fill=orangeFill
                                            for pe in range(int(d)+1,col_dict_rem['KPI Next 2 Days Validated']+1):
                                                ws[str(chr(65+int(pe)))+str(j)]='Pending'
                                                ws[str(chr(65+int(pe)))+str(j)].fill=goldFill
                                                
                        ll=str(ws[str(chr((col_dict_rem['Locked Tech'])+65))+str(j)].value).split(', ') #locked list
                        otj='' #overdue tech join
                        for il in ll:
                            fp=il.split('_')[0] #fp is first part
                            if fp=='NR2500' or fp=='NR600':
                                otj=', '.join([otj,il])
                        otj=otj.strip(', ')
                        if otj=='None' or otj==None or otj=='':
                            break
                        else:
                            if (ws[nextcol2+str(j)].value=='Site ID not found') or (ws[nextcol2+str(j)].value=='') or (ws[nextcol2+str(j)].value=='None'):
                                ws[nextcol2+str(j)]='KPI validation is pending as New Tech '+otj+' is locked'
                            else:
                                ws[nextcol2+str(j)]=ws[nextcol2+str(j)].value+'\n'+'KPI validation is pending as New Tech '+otj+' is locked'
            
            
            final_output=generate_5g_remarks(site_list,dataset3,col_dict_df)
                
            #print(final_output)
            df = pd.DataFrame.from_dict(final_output, orient="Index")
            k=''
            for i in list(df.index):
                for j in range(2,trows1+1):
                    if i == ws['A'+str(j)].value:
                        if (ws[nextcol5+str(j)].value=='Site ID not found') or (ws[nextcol5+str(j)].value=='') or (ws[nextcol5+str(j)].value=='None'): # or (ws[nextcol5+str(j)].value.find('issue was observed')!=-1):
                            ws[nextcol5+str(j)] = str((df.loc[i,0]))
                        else:
                            ws[nextcol5+str(j)] = ws[nextcol5+str(j)].value+'\n'+str((df.loc[i,0]))
                        if str((df.loc[i,0]))=='':
                            for a,b in duration_meaning.items():
                                if int(b)==file_duration:
                                    k=a
                                    for c,d in col_dict_rem.items():
                                        if k==c:
                                            if not ws[chr(col_dict_rem['Locked Tech']+65)+str(j)].value: #check if blank or None
                                            # ws[str(chr(65+int(d)))+str(j)]='Pass'
                                            # ws[str(chr(65+int(d)))+str(j)].fill=neongreenFill
                                                if (ws[nextcol5+str(j)].value=='Site ID not found') or (ws[nextcol5+str(j)].value=='') or (ws[nextcol5+str(j)].value=='None'):
                                                    ws[nextcol5+str(j)]='During '+str(ws[str(chr((col_dict_rem['Hour'])+65))+str(j)].value)+' hour KPI validation, no 5G issue was observed'
                                                else:
                                                    ws[nextcol5+str(j)]=ws[nextcol5+str(j)].value+'\n'+'During '+str(ws[str(chr((col_dict_rem['Hour'])+65))+str(j)].value)+' hour KPI validation, no 5G issue was observed'
                                           # else:
                                               # ws[str(chr(65+int(d)))+str(j)]='Overdue'
                                               # ws[str(chr(65+int(d)))+str(j)].fill=orangeFill
                                               # ws[nextcol5+str(j)]='KPI validation is pending as New Tech '+str(ws[str(chr((col_dict_rem['Locked Tech'])+65))+str(j)].value)+' is locked/down'
                                           # for pe in range(int(d)+1,col_dict_rem['KPI Next 2 Days Validated']+1):
                                           #     ws[str(chr(65+int(pe)))+str(j)]='Pending'
                                           #     ws[str(chr(65+int(pe)))+str(j)].fill=goldFill
                        # elif str((df.loc[i,0]))!='':
                        #     for a,b in duration_meaning.items():
                        #         if int(b)==file_duration:
                        #             k=a
                        #             for c,d in col_dict_rem.items():
                        #                if k==c:
                        #                    if not ws[chr(col_dict_rem['Locked Tech']+65)+str(j)].value:
                        #                        ws[str(chr(65+int(d)))+str(j)]='Fail'
                        #                        ws[str(chr(65+int(d)))+str(j)].fill=redFill
                        #                    else:
                        #                        ws[str(chr(65+int(d)))+str(j)]='Overdue'
                        #                        ws[str(chr(65+int(d)))+str(j)].fill=orangeFill
                        #                    for pe in range(int(d)+1,col_dict_rem['KPI Next 2 Days Validated']+1):
                        #                        ws[str(chr(65+int(pe)))+str(j)]='Pending'
                        #                        ws[str(chr(65+int(pe)))+str(j)].fill=goldFill
                        # break
                    
                # for k in range(2,trows2+1):
                #     if i == ws2['A'+str(k)].value:
                #         ws2[nextcol5+str(k)] = str((df.loc[i,0]))
                
            final_output=generate_5g_remarks(site_list,dataset,col_dict_df)
                
            #print(final_output)
            df = pd.DataFrame.from_dict(final_output, orient="Index")
            k=''
            for i in list(df.index):
                for j in range(2,trows1+1):
                    if i == ws['A'+str(j)].value:
                        if (ws[nextcol6+str(j)].value=='Site ID not found') or (ws[nextcol6+str(j)].value=='') or (ws[nextcol6+str(j)].value=='None'): # or (ws[nextcol6+str(j)].value.find('issue was observed')!=-1):
                            ws[nextcol6+str(j)] = str((df.loc[i,0]))
                        else:
                            ws[nextcol6+str(j)] = ws[nextcol6+str(j)].value+'\n'+str((df.loc[i,0]))
                        if str((df.loc[i,0]))=='':
                            for a,b in duration_meaning.items():
                                if int(b)==file_duration:
                                    k=a
                                    for c,d in col_dict_rem.items():
                                        if k==c:
                                            if not ws[chr(col_dict_rem['Locked Tech']+65)+str(j)].value: #check if blank or None
                                                # ws[str(chr(65+int(d)))+str(j)]='Pass'
                                                # ws[str(chr(65+int(d)))+str(j)].fill=neongreenFill
                                                if (ws[nextcol6+str(j)].value=='Site ID not found') or (ws[nextcol6+str(j)].value=='') or (ws[nextcol6+str(j)].value=='None'):
                                                    ws[nextcol6+str(j)]='During '+str(ws[str(chr((col_dict_rem['Hour'])+65))+str(j)].value)+' hour KPI validation, no 5G issue was observed'
                                                else:
                                                    ws[nextcol6+str(j)]=ws[nextcol6+str(j)].value+'\n'+'During '+str(ws[str(chr((col_dict_rem['Hour'])+65))+str(j)].value)+' hour KPI validation, no 5G issue was observed'
                                               # else:
                                                   # ws[str(chr(65+int(d)))+str(j)]='Overdue'
                                                   # ws[str(chr(65+int(d)))+str(j)].fill=orangeFill
                                                   # ws[nextcol6+str(j)]='KPI validation is pending as New Tech '+str(ws[str(chr((col_dict_rem['Locked Tech'])+65))+str(j)].value)+' is locked/down'
                                               # for pe in range(int(d)+1,col_dict_rem['KPI Next 2 Days Validated']+1):
                                               #     ws[str(chr(65+int(pe)))+str(j)]='Pending'
                                               #     ws[str(chr(65+int(pe)))+str(j)].fill=goldFill
                            # elif str((df.loc[i,0]))!='':
                            #     for a,b in duration_meaning.items():
                            #         if int(b)==file_duration:
                            #             k=a
                            #             for c,d in col_dict_rem.items():
                            #                if k==c:
                            #                    if not ws[chr(col_dict_rem['Locked Tech']+65)+str(j)].value:
                            #                        ws[str(chr(65+int(d)))+str(j)]='Fail'
                            #                        ws[str(chr(65+int(d)))+str(j)].fill=redFill
                            #                    else:
                            #                        ws[str(chr(65+int(d)))+str(j)]='Overdue'
                            #                        ws[str(chr(65+int(d)))+str(j)].fill=orangeFill
                            #                    for pe in range(int(d)+1,col_dict_rem['KPI Next 2 Days Validated']+1):
                            #                        ws[str(chr(65+int(pe)))+str(j)]='Pending'
                            #                        ws[str(chr(65+int(pe)))+str(j)].fill=goldFill
                            # break
                    
                # for k in range(2,trows2+1):
                #     if i == ws2['A'+str(k)].value:
                #         ws2[nextcol5+str(k)] = str((df.loc[i,0]))
                #         break
            
        if (file_name.find('GSM')!=-1):
        
            final_output=generate_gsm_remarks(site_list,dataset2,col_dict_df)
                
            #print(final_output)
            df = pd.DataFrame.from_dict(final_output, orient="Index")
            k=''
            for i in list(df.index):
                for j in range(2,trows1+1):
                    if i == ws['A'+str(j)].value:
                        if (str((df.loc[i,0]))!='') and (ws[nextcol3+str(j)].value=='Site ID not found') or (ws[nextcol3+str(j)].value=='') or (ws[nextcol3+str(j)].value=='None'): # or (ws[nextcol3+str(j)].value.find('issue was observed')!=-1):
                            ws[nextcol3+str(j)] = 'During '+str(ws[str(chr((col_dict_rem['Hour'])+65))+str(j)].value)+' hour KPI validation, the following issues were observed:'
                            ws[nextcol3+str(j)] = ws[nextcol3+str(j)].value+'\n'+str((df.loc[i,0]))
                        if str((df.loc[i,0]))=='':
                            for a,b in duration_meaning.items():
                                if int(b)==file_duration:
                                    k=a
                                    for c,d in col_dict_rem.items():
                                        if k==c:
                                            if not ws[chr(col_dict_rem['Locked Tech']+65)+str(j)].value: #check if blank or None
                                                # ws[str(chr(65+int(d)))+str(j)]='Pass'
                                                # ws[str(chr(65+int(d)))+str(j)].fill=neongreenFill
                                                if (ws[nextcol3+str(j)].value=='Site ID not found') or (ws[nextcol3+str(j)].value=='') or (ws[nextcol3+str(j)].value=='None'):
                                                    ws[nextcol3+str(j)]='During '+str(ws[str(chr((col_dict_rem['Hour'])+65))+str(j)].value)+' hour KPI validation, no GSM issue was observed'
                                                    if not ws[str(chr(65+int(d)))+str(j)].value or ws[str(chr(65+int(d)))+str(j)].fill.start_color.index=='00000000':
                                                        ws[str(chr(65+int(d)))+str(j)]='Pass'
                                                        ws[str(chr(65+int(d)))+str(j)].fill=neongreenFill
                                                # elif (ws[nextcol3+str(j)].value.find('issue was observed')!=-1):
                                                #         ws[str(chr(65+int(d)))+str(j)]='Pass'
                                                #         ws[str(chr(65+int(d)))+str(j)].fill=neongreenFill
                                                #         ws[nextcol3+str(j)]=ws[nextcol3+str(j)].value+'During '+str(ws[str(chr((col_dict_rem['Hour'])+65))+str(j)].value)+' hour KPI validation, no GSM issue was observed'
                                                # else:
                                                #     ws[str(chr(65+int(d)))+str(j)]='Fail'
                                                #     ws[str(chr(65+int(d)))+str(j)].fill=redFill
                                                #     ws[nextcol3+str(j)]=ws[nextcol3+str(j)].value+'\n'+'During '+str(ws[str(chr((col_dict_rem['Hour'])+65))+str(j)].value)+' hour KPI validation, no GSM issue was observed'
                                            else:
                                                if (ws[nextcol3+str(j)].value=='Site ID not found') or (ws[nextcol3+str(j)].value=='') or (ws[nextcol3+str(j)].value=='None'):
                                                    ws[nextcol3+str(j)]='During '+str(ws[str(chr((col_dict_rem['Hour'])+65))+str(j)].value)+' hour KPI validation, no GSM issue was observed'
                                                else:
                                                    ws[nextcol3+str(j)]=ws[nextcol3+str(j)].value+'\n'+'During '+str(ws[str(chr((col_dict_rem['Hour'])+65))+str(j)].value)+' hour KPI validation, no GSM issue was observed'
                                                ws[str(chr(65+int(d)))+str(j)]='Overdue'
                                                ws[str(chr(65+int(d)))+str(j)].fill=orangeFill 
                                                # ll=str(ws[str(chr((col_dict_rem['Locked Tech'])+65))+str(j)].value).split(', ') #locked list
                                                # otj='' #overdue tech join
                                                # for il in ll:
                                                #     if il!='NR2500' and il!='NR600' and il!='GSM' and il!='U1900' and il!='U2100':
                                                #         otj=', '.join([otj,il])
                                                # otj=otj.strip(', ')
                                                # if not otj:
                                                #     break
                                                # else:
                                                #     if (ws[nextcol3+str(j)].value=='Site ID not found') or (ws[nextcol3+str(j)].value=='') or (ws[nextcol3+str(j)].value=='None'):
                                                #         ws[nextcol3+str(j)]='KPI validation is pending as New Tech '+otj+' is locked'
                                                #     else:
                                                #         ws[nextcol3+str(j)]=ws[nextcol3+str(j)].value+'\n'+'KPI validation is pending as New Tech '+otj+' is locked'
                                            for pe in range(int(d)+1,col_dict_rem['KPI Next 2 Days Validated']+1):
                                                ws[str(chr(65+int(pe)))+str(j)]='Pending'
                                                ws[str(chr(65+int(pe)))+str(j)].fill=goldFill
                        # ll=str(ws[str(chr((col_dict_rem['Locked Tech'])+65))+str(j)].value).split(', ') #locked list
                        # otj='' #overdue tech join
                        # for il in ll:
                        #     if il!='NR2500' and il!='NR600' and il!='GSM' and il!='U1900' and il!='U2100':
                        #         otj=', '.join([otj,il])
                        #         otj=otj.strip(', ')
                        #         if not otj:
                        #             break
                        #         else:
                        #             if (ws[nextcol3+str(j)].value=='Site ID not found') or (ws[nextcol3+str(j)].value=='') or (ws[nextcol3+str(j)].value=='None'):
                        #                 ws[nextcol3+str(j)]='KPI validation is pending as New Tech '+otj+' is locked'
                        #             else:
                        #                 ws[nextcol3+str(j)]=ws[nextcol3+str(j)].value+'\n'+'KPI validation is pending as New Tech '+otj+' is locked'
                                                
                                                   
                        elif str((df.loc[i,0]))!='':
                            for a,b in duration_meaning.items():
                                if int(b)==file_duration:
                                    k=a
                                    for c,d in col_dict_rem.items():
                                        if k==c:
                                            if not ws[chr(col_dict_rem['Locked Tech']+65)+str(j)].value:
                                                ws[str(chr(65+int(d)))+str(j)]='Fail'
                                                ws[str(chr(65+int(d)))+str(j)].fill=redFill
                                            else:
                                                ws[str(chr(65+int(d)))+str(j)]='Overdue'
                                                ws[str(chr(65+int(d)))+str(j)].fill=orangeFill
                                            for pe in range(int(d)+1,col_dict_rem['KPI Next 2 Days Validated']+1):
                                                ws[str(chr(65+int(pe)))+str(j)]='Pending'
                                                ws[str(chr(65+int(pe)))+str(j)].fill=goldFill
                                                
                        ll=str(ws[str(chr((col_dict_rem['Locked Tech'])+65))+str(j)].value).split(', ') #locked list
                        otj='' #overdue tech join
                        for il in ll:
                            fp=il.split('_')[0] #fp is first part
                            if fp=='GSM':
                                otj=', '.join([otj,il])
                        otj=otj.strip(', ')
                        if otj=='None' or otj==None or otj=='':
                            break
                        else:
                            if (ws[nextcol3+str(j)].value=='Site ID not found') or (ws[nextcol3+str(j)].value=='') or (ws[nextcol3+str(j)].value=='None'):
                                ws[nextcol3+str(j)]='KPI validation is pending as New Tech '+otj+' is locked'
                            else:
                                ws[nextcol3+str(j)]=ws[nextcol3+str(j)].value+'\n'+'KPI validation is pending as New Tech '+otj+' is locked'
            
            
            final_output=generate_gsm_remarks(site_list,dataset3,col_dict_df)
                
            #print(final_output)
            df = pd.DataFrame.from_dict(final_output, orient="Index")
            k=''
            for i in list(df.index):
                for j in range(2,trows1+1):
                    if i == ws['A'+str(j)].value:
                        if (ws[nextcol5+str(j)].value=='Site ID not found') or (ws[nextcol5+str(j)].value=='') or (ws[nextcol5+str(j)].value=='None'): # or (ws[nextcol5+str(j)].value.find('issue was observed')!=-1):
                            ws[nextcol5+str(j)] = str((df.loc[i,0]))
                        else:
                            ws[nextcol5+str(j)] = ws[nextcol5+str(j)].value+'\n'+str((df.loc[i,0]))
                        if str((df.loc[i,0]))=='':
                            for a,b in duration_meaning.items():
                                if int(b)==file_duration:
                                    k=a
                                    for c,d in col_dict_rem.items():
                                        if k==c:
                                            if not ws[chr(col_dict_rem['Locked Tech']+65)+str(j)].value: #check if blank or None
                                            # ws[str(chr(65+int(d)))+str(j)]='Pass'
                                            # ws[str(chr(65+int(d)))+str(j)].fill=neongreenFill
                                                if (ws[nextcol5+str(j)].value=='Site ID not found') or (ws[nextcol5+str(j)].value=='') or (ws[nextcol5+str(j)].value=='None'):
                                                    ws[nextcol5+str(j)]='During '+str(ws[str(chr((col_dict_rem['Hour'])+65))+str(j)].value)+' hour KPI validation, no GSM issue was observed'
                                                else:
                                                    ws[nextcol5+str(j)]=ws[nextcol5+str(j)].value+'\n'+'During '+str(ws[str(chr((col_dict_rem['Hour'])+65))+str(j)].value)+' hour KPI validation, no GSM issue was observed'
                                           # else:
                                               # ws[str(chr(65+int(d)))+str(j)]='Overdue'
                                               # ws[str(chr(65+int(d)))+str(j)].fill=orangeFill
                                               # ws[nextcol5+str(j)]='KPI validation is pending as New Tech '+str(ws[str(chr((col_dict_rem['Locked Tech'])+65))+str(j)].value)+' is locked/down'
                                           # for pe in range(int(d)+1,col_dict_rem['KPI Next 2 Days Validated']+1):
                                           #     ws[str(chr(65+int(pe)))+str(j)]='Pending'
                                           #     ws[str(chr(65+int(pe)))+str(j)].fill=goldFill
                        # elif str((df.loc[i,0]))!='':
                        #     for a,b in duration_meaning.items():
                        #         if int(b)==file_duration:
                        #             k=a
                        #             for c,d in col_dict_rem.items():
                        #                if k==c:
                        #                    if not ws[chr(col_dict_rem['Locked Tech']+65)+str(j)].value:
                        #                        ws[str(chr(65+int(d)))+str(j)]='Fail'
                        #                        ws[str(chr(65+int(d)))+str(j)].fill=redFill
                        #                    else:
                        #                        ws[str(chr(65+int(d)))+str(j)]='Overdue'
                        #                        ws[str(chr(65+int(d)))+str(j)].fill=orangeFill
                        #                    for pe in range(int(d)+1,col_dict_rem['KPI Next 2 Days Validated']+1):
                        #                        ws[str(chr(65+int(pe)))+str(j)]='Pending'
                        #                        ws[str(chr(65+int(pe)))+str(j)].fill=goldFill
                        # break
                    
                # for k in range(2,trows2+1):
                #     if i == ws2['A'+str(k)].value:
                #         ws2[nextcol3+str(k)] = str((df.loc[i,0]))
                
            final_output=generate_gsm_remarks(site_list,dataset,col_dict_df)
                
            #print(final_output)
            df = pd.DataFrame.from_dict(final_output, orient="Index")
            k=''
            for i in list(df.index):
                for j in range(2,trows1+1):
                    if i == ws['A'+str(j)].value:
                        if (ws[nextcol6+str(j)].value=='Site ID not found') or (ws[nextcol6+str(j)].value=='') or (ws[nextcol6+str(j)].value=='None'): # or (ws[nextcol6+str(j)].value.find('issue was observed')!=-1):
                            ws[nextcol6+str(j)] = str((df.loc[i,0]))
                        else:
                            ws[nextcol6+str(j)] = ws[nextcol6+str(j)].value+'\n'+str((df.loc[i,0]))
                        if str((df.loc[i,0]))=='':
                            for a,b in duration_meaning.items():
                                if int(b)==file_duration:
                                    k=a
                                    for c,d in col_dict_rem.items():
                                        if k==c:
                                            if not ws[chr(col_dict_rem['Locked Tech']+65)+str(j)].value: #check if blank or None
                                                # ws[str(chr(65+int(d)))+str(j)]='Pass'
                                                # ws[str(chr(65+int(d)))+str(j)].fill=neongreenFill
                                                if (ws[nextcol6+str(j)].value=='Site ID not found') or (ws[nextcol6+str(j)].value=='') or (ws[nextcol6+str(j)].value=='None'):
                                                    ws[nextcol6+str(j)]='During '+str(ws[str(chr((col_dict_rem['Hour'])+65))+str(j)].value)+' hour KPI validation, no GSM issue was observed'
                                                else:
                                                    ws[nextcol6+str(j)]=ws[nextcol6+str(j)].value+'\n'+'During '+str(ws[str(chr((col_dict_rem['Hour'])+65))+str(j)].value)+' hour KPI validation, no GSM issue was observed'
                                               # else:
                                                   # ws[str(chr(65+int(d)))+str(j)]='Overdue'
                                                   # ws[str(chr(65+int(d)))+str(j)].fill=orangeFill
                                                   # ws[nextcol6+str(j)]='KPI validation is pending as New Tech '+str(ws[str(chr((col_dict_rem['Locked Tech'])+65))+str(j)].value)+' is locked/down'
                                               # for pe in range(int(d)+1,col_dict_rem['KPI Next 2 Days Validated']+1):
                                               #     ws[str(chr(65+int(pe)))+str(j)]='Pending'
                                               #     ws[str(chr(65+int(pe)))+str(j)].fill=goldFill
                            # elif str((df.loc[i,0]))!='':
                            #     for a,b in duration_meaning.items():
                            #         if int(b)==file_duration:
                            #             k=a
                            #             for c,d in col_dict_rem.items():
                            #                if k==c:
                            #                    if not ws[chr(col_dict_rem['Locked Tech']+65)+str(j)].value:
                            #                        ws[str(chr(65+int(d)))+str(j)]='Fail'
                            #                        ws[str(chr(65+int(d)))+str(j)].fill=redFill
                            #                    else:
                            #                        ws[str(chr(65+int(d)))+str(j)]='Overdue'
                            #                        ws[str(chr(65+int(d)))+str(j)].fill=orangeFill
                            #                    for pe in range(int(d)+1,col_dict_rem['KPI Next 2 Days Validated']+1):
                            #                        ws[str(chr(65+int(pe)))+str(j)]='Pending'
                            #                        ws[str(chr(65+int(pe)))+str(j)].fill=goldFill
                            # break
                    
                # for k in range(2,trows2+1):
                #     if i == ws2['A'+str(k)].value:
                #         ws2[nextcol5+str(k)] = str((df.loc[i,0]))
                #         break
            
        if (file_name.find('UMTS')!=-1):
        
            final_output=generate_umts_remarks(site_list,dataset2,col_dict_df)
                
            #print(final_output)
            df = pd.DataFrame.from_dict(final_output, orient="Index")
            k=''
            for i in list(df.index):
                for j in range(2,trows1+1):
                    if i == ws['A'+str(j)].value:
                        if (str((df.loc[i,0]))!='') and (ws[nextcol4+str(j)].value=='Site ID not found') or (ws[nextcol4+str(j)].value=='') or (ws[nextcol4+str(j)].value=='None'): # or (ws[nextcol4+str(j)].value.find('issue was observed')!=-1):
                            ws[nextcol4+str(j)] = 'During '+str(ws[str(chr((col_dict_rem['Hour'])+65))+str(j)].value)+' hour KPI validation, the following issues were observed:'
                            ws[nextcol4+str(j)] = ws[nextcol4+str(j)].value+'\n'+str((df.loc[i,0]))
                        if str((df.loc[i,0]))=='':
                            for a,b in duration_meaning.items():
                                if int(b)==file_duration:
                                    k=a
                                    for c,d in col_dict_rem.items():
                                        if k==c:
                                            if not ws[chr(col_dict_rem['Locked Tech']+65)+str(j)].value: #check if blank or None
                                                # ws[str(chr(65+int(d)))+str(j)]='Pass'
                                                # ws[str(chr(65+int(d)))+str(j)].fill=neongreenFill
                                                if (ws[nextcol4+str(j)].value=='Site ID not found') or (ws[nextcol4+str(j)].value=='') or (ws[nextcol4+str(j)].value=='None'):
                                                    ws[nextcol4+str(j)]='During '+str(ws[str(chr((col_dict_rem['Hour'])+65))+str(j)].value)+' hour KPI validation, no UMTS issue was observed'
                                                    if not ws[str(chr(65+int(d)))+str(j)].value or ws[str(chr(65+int(d)))+str(j)].fill.start_color.index=='00000000':
                                                        ws[str(chr(65+int(d)))+str(j)]='Pass'
                                                        ws[str(chr(65+int(d)))+str(j)].fill=neongreenFill
                                                # elif (ws[nextcol4+str(j)].value.find('issue was observed')!=-1):
                                                #         ws[str(chr(65+int(d)))+str(j)]='Pass'
                                                #         ws[str(chr(65+int(d)))+str(j)].fill=neongreenFill
                                                #         ws[nextcol4+str(j)]=ws[nextcol4+str(j)].value+'During '+str(ws[str(chr((col_dict_rem['Hour'])+65))+str(j)].value)+' hour KPI validation, no UMTS issue was observed'
                                                # else:
                                                #     ws[str(chr(65+int(d)))+str(j)]='Fail'
                                                #     ws[str(chr(65+int(d)))+str(j)].fill=redFill
                                                #     ws[nextcol4+str(j)]=ws[nextcol4+str(j)].value+'\n'+'During '+str(ws[str(chr((col_dict_rem['Hour'])+65))+str(j)].value)+' hour KPI validation, no UMTS issue was observed'
                                            else:
                                                if (ws[nextcol4+str(j)].value=='Site ID not found') or (ws[nextcol4+str(j)].value=='') or (ws[nextcol4+str(j)].value=='None'):
                                                    ws[nextcol4+str(j)]='During '+str(ws[str(chr((col_dict_rem['Hour'])+65))+str(j)].value)+' hour KPI validation, no UMTS issue was observed'
                                                else:
                                                    ws[nextcol4+str(j)]=ws[nextcol4+str(j)].value+'\n'+'During '+str(ws[str(chr((col_dict_rem['Hour'])+65))+str(j)].value)+' hour KPI validation, no UMTS issue was observed'
                                                ws[str(chr(65+int(d)))+str(j)]='Overdue'
                                                ws[str(chr(65+int(d)))+str(j)].fill=orangeFill 
                                                # ll=str(ws[str(chr((col_dict_rem['Locked Tech'])+65))+str(j)].value).split(', ') #locked list
                                                # otj='' #overdue tech join
                                                # for il in ll:
                                                #     if il!='NR2500' and il!='NR600' and il!='UMTS' and il!='U1900' and il!='U2100':
                                                #         otj=', '.join([otj,il])
                                                # otj=otj.strip(', ')
                                                # if not otj:
                                                #     break
                                                # else:
                                                #     if (ws[nextcol4+str(j)].value=='Site ID not found') or (ws[nextcol4+str(j)].value=='') or (ws[nextcol4+str(j)].value=='None'):
                                                #         ws[nextcol4+str(j)]='KPI validation is pending as New Tech '+otj+' is locked'
                                                #     else:
                                                #         ws[nextcol4+str(j)]=ws[nextcol4+str(j)].value+'\n'+'KPI validation is pending as New Tech '+otj+' is locked'
                                            for pe in range(int(d)+1,col_dict_rem['KPI Next 2 Days Validated']+1):
                                                ws[str(chr(65+int(pe)))+str(j)]='Pending'
                                                ws[str(chr(65+int(pe)))+str(j)].fill=goldFill
                        # ll=str(ws[str(chr((col_dict_rem['Locked Tech'])+65))+str(j)].value).split(', ') #locked list
                        # otj='' #overdue tech join
                        # for il in ll:
                        #     if il!='NR2500' and il!='NR600' and il!='UMTS' and il!='U1900' and il!='U2100':
                        #         otj=', '.join([otj,il])
                        #         otj=otj.strip(', ')
                        #         if not otj:
                        #             break
                        #         else:
                        #             if (ws[nextcol4+str(j)].value=='Site ID not found') or (ws[nextcol4+str(j)].value=='') or (ws[nextcol4+str(j)].value=='None'):
                        #                 ws[nextcol4+str(j)]='KPI validation is pending as New Tech '+otj+' is locked'
                        #             else:
                        #                 ws[nextcol4+str(j)]=ws[nextcol4+str(j)].value+'\n'+'KPI validation is pending as New Tech '+otj+' is locked'
                                                
                                                   
                        elif str((df.loc[i,0]))!='':
                            for a,b in duration_meaning.items():
                                if int(b)==file_duration:
                                    k=a
                                    for c,d in col_dict_rem.items():
                                        if k==c:
                                            if not ws[chr(col_dict_rem['Locked Tech']+65)+str(j)].value:
                                                ws[str(chr(65+int(d)))+str(j)]='Fail'
                                                ws[str(chr(65+int(d)))+str(j)].fill=redFill
                                            else:
                                                ws[str(chr(65+int(d)))+str(j)]='Overdue'
                                                ws[str(chr(65+int(d)))+str(j)].fill=orangeFill
                                            for pe in range(int(d)+1,col_dict_rem['KPI Next 2 Days Validated']+1):
                                                ws[str(chr(65+int(pe)))+str(j)]='Pending'
                                                ws[str(chr(65+int(pe)))+str(j)].fill=goldFill
                                                
                        ll=str(ws[str(chr((col_dict_rem['Locked Tech'])+65))+str(j)].value).split(', ') #locked list
                        otj='' #overdue tech join
                        for il in ll:
                            fp=il.split('_')[0] #fp is first part
                            if fp=='U1900' or fp=='U2100':
                                otj=', '.join([otj,il])
                        otj=otj.strip(', ')
                        if otj=='None' or otj==None or otj=='':
                            break
                        else:
                            if (ws[nextcol4+str(j)].value=='Site ID not found') or (ws[nextcol4+str(j)].value=='') or (ws[nextcol4+str(j)].value=='None'):
                                ws[nextcol4+str(j)]='KPI validation is pending as New Tech '+otj+' is locked'
                            else:
                                ws[nextcol4+str(j)]=ws[nextcol4+str(j)].value+'\n'+'KPI validation is pending as New Tech '+otj+' is locked'
            
            
            final_output=generate_umts_remarks(site_list,dataset3,col_dict_df)
                
            #print(final_output)
            df = pd.DataFrame.from_dict(final_output, orient="Index")
            k=''
            for i in list(df.index):
                for j in range(2,trows1+1):
                    if i == ws['A'+str(j)].value:
                        if (ws[nextcol5+str(j)].value=='Site ID not found') or (ws[nextcol5+str(j)].value=='') or (ws[nextcol5+str(j)].value=='None'): # or (ws[nextcol5+str(j)].value.find('issue was observed')!=-1):
                            ws[nextcol5+str(j)] = str((df.loc[i,0]))
                        else:
                            ws[nextcol5+str(j)] = ws[nextcol5+str(j)].value+'\n'+str((df.loc[i,0]))
                        if str((df.loc[i,0]))=='':
                            for a,b in duration_meaning.items():
                                if int(b)==file_duration:
                                    k=a
                                    for c,d in col_dict_rem.items():
                                        if k==c:
                                            if not ws[chr(col_dict_rem['Locked Tech']+65)+str(j)].value: #check if blank or None
                                            # ws[str(chr(65+int(d)))+str(j)]='Pass'
                                            # ws[str(chr(65+int(d)))+str(j)].fill=neongreenFill
                                                if (ws[nextcol5+str(j)].value=='Site ID not found') or (ws[nextcol5+str(j)].value=='') or (ws[nextcol5+str(j)].value=='None'):
                                                    ws[nextcol5+str(j)]='During '+str(ws[str(chr((col_dict_rem['Hour'])+65))+str(j)].value)+' hour KPI validation, no UMTS issue was observed'
                                                else:
                                                    ws[nextcol5+str(j)]=ws[nextcol5+str(j)].value+'\n'+'During '+str(ws[str(chr((col_dict_rem['Hour'])+65))+str(j)].value)+' hour KPI validation, no UMTS issue was observed'
                                           # else:
                                               # ws[str(chr(65+int(d)))+str(j)]='Overdue'
                                               # ws[str(chr(65+int(d)))+str(j)].fill=orangeFill
                                               # ws[nextcol5+str(j)]='KPI validation is pending as New Tech '+str(ws[str(chr((col_dict_rem['Locked Tech'])+65))+str(j)].value)+' is locked/down'
                                           # for pe in range(int(d)+1,col_dict_rem['KPI Next 2 Days Validated']+1):
                                           #     ws[str(chr(65+int(pe)))+str(j)]='Pending'
                                           #     ws[str(chr(65+int(pe)))+str(j)].fill=goldFill
                        # elif str((df.loc[i,0]))!='':
                        #     for a,b in duration_meaning.items():
                        #         if int(b)==file_duration:
                        #             k=a
                        #             for c,d in col_dict_rem.items():
                        #                if k==c:
                        #                    if not ws[chr(col_dict_rem['Locked Tech']+65)+str(j)].value:
                        #                        ws[str(chr(65+int(d)))+str(j)]='Fail'
                        #                        ws[str(chr(65+int(d)))+str(j)].fill=redFill
                        #                    else:
                        #                        ws[str(chr(65+int(d)))+str(j)]='Overdue'
                        #                        ws[str(chr(65+int(d)))+str(j)].fill=orangeFill
                        #                    for pe in range(int(d)+1,col_dict_rem['KPI Next 2 Days Validated']+1):
                        #                        ws[str(chr(65+int(pe)))+str(j)]='Pending'
                        #                        ws[str(chr(65+int(pe)))+str(j)].fill=goldFill
                        # break
                    
                # for k in range(2,trows2+1):
                #     if i == ws2['A'+str(k)].value:
                #         ws2[nextcol4+str(k)] = str((df.loc[i,0]))
                
            final_output=generate_umts_remarks(site_list,dataset,col_dict_df)
                
            #print(final_output)
            df = pd.DataFrame.from_dict(final_output, orient="Index")
            k=''
            for i in list(df.index):
                for j in range(2,trows1+1):
                    if i == ws['A'+str(j)].value:
                        if (ws[nextcol6+str(j)].value=='Site ID not found') or (ws[nextcol6+str(j)].value=='') or (ws[nextcol6+str(j)].value=='None'): # or (ws[nextcol6+str(j)].value.find('issue was observed')!=-1):
                            ws[nextcol6+str(j)] = str((df.loc[i,0]))
                        else:
                            ws[nextcol6+str(j)] = ws[nextcol6+str(j)].value+'\n'+str((df.loc[i,0]))
                        if str((df.loc[i,0]))=='':
                            for a,b in duration_meaning.items():
                                if int(b)==file_duration:
                                    k=a
                                    for c,d in col_dict_rem.items():
                                        if k==c:
                                            if not ws[chr(col_dict_rem['Locked Tech']+65)+str(j)].value: #check if blank or None
                                                # ws[str(chr(65+int(d)))+str(j)]='Pass'
                                                # ws[str(chr(65+int(d)))+str(j)].fill=neongreenFill
                                                if (ws[nextcol6+str(j)].value=='Site ID not found') or (ws[nextcol6+str(j)].value=='') or (ws[nextcol6+str(j)].value=='None'):
                                                    ws[nextcol6+str(j)]='During '+str(ws[str(chr((col_dict_rem['Hour'])+65))+str(j)].value)+' hour KPI validation, no UMTS issue was observed'
                                                else:
                                                    ws[nextcol6+str(j)]=ws[nextcol6+str(j)].value+'\n'+'During '+str(ws[str(chr((col_dict_rem['Hour'])+65))+str(j)].value)+' hour KPI validation, no UMTS issue was observed'
                                               # else:
                                                   # ws[str(chr(65+int(d)))+str(j)]='Overdue'
                                                   # ws[str(chr(65+int(d)))+str(j)].fill=orangeFill
                                                   # ws[nextcol6+str(j)]='KPI validation is pending as New Tech '+str(ws[str(chr((col_dict_rem['Locked Tech'])+65))+str(j)].value)+' is locked/down'
                                               # for pe in range(int(d)+1,col_dict_rem['KPI Next 2 Days Validated']+1):
                                               #     ws[str(chr(65+int(pe)))+str(j)]='Pending'
                                               #     ws[str(chr(65+int(pe)))+str(j)].fill=goldFill
                            # elif str((df.loc[i,0]))!='':
                            #     for a,b in duration_meaning.items():
                            #         if int(b)==file_duration:
                            #             k=a
                            #             for c,d in col_dict_rem.items():
                            #                if k==c:
                            #                    if not ws[chr(col_dict_rem['Locked Tech']+65)+str(j)].value:
                            #                        ws[str(chr(65+int(d)))+str(j)]='Fail'
                            #                        ws[str(chr(65+int(d)))+str(j)].fill=redFill
                            #                    else:
                            #                        ws[str(chr(65+int(d)))+str(j)]='Overdue'
                            #                        ws[str(chr(65+int(d)))+str(j)].fill=orangeFill
                            #                    for pe in range(int(d)+1,col_dict_rem['KPI Next 2 Days Validated']+1):
                            #                        ws[str(chr(65+int(pe)))+str(j)]='Pending'
                            #                        ws[str(chr(65+int(pe)))+str(j)].fill=goldFill
                            # break
                    
                # for k in range(2,trows2+1):
                #     if i == ws2['A'+str(k)].value:
                #         ws2[nextcol5+str(k)] = str((df.loc[i,0]))
                #         break
            

for j in range(2,trows1+1):
    existstr=ws[nextcol5+str(j)].value
    existstr_list=existstr.split('\n')
    for item in existstr_list:
        if item=='' or item==None:
            existstr_list.remove(item)
    k=1
    for i in range(0,len(existstr_list)):
        if len(existstr_list[i])<=2 or existstr_list[i]==None:
            break
        if existstr_list[i][0].isdigit()==1 and existstr_list[i][1:3]=='. ':
            existstr_list[i]=str(k)+existstr_list[i][1:]
            k+=1
    for i in range(0,len(existstr_list)):
        if len(existstr_list[i])<=2 or existstr_list[i]==None:
            break
        if existstr_list[i][0:3]=='1. ':
            existstr_list.insert(i,'During '+str(ws[str(chr((col_dict_rem['Hour'])+65))+str(j)].value)+' hour KPI validation, the following issues were observed:')
            break
    ws[nextcol5+str(j)]='\n'.join(existstr_list)

for j in range(2,trows1+1):
    newstr=ws[nextcol6+str(j)].value
    newstr_list=newstr.split('\n')
    for item in newstr_list:
        if item=='' or item==None:
            newstr_list.remove(item)
    k=1
    for i in range(0,len(newstr_list)):
        if len(newstr_list[i])<=2 or newstr_list[i]==None:
            break
        if newstr_list[i][0].isdigit()==1 and newstr_list[i][1:3]=='. ':
            newstr_list[i]=str(k)+newstr_list[i][1:]
            k+=1                                                                                                                                                                                                                                                     
    for i in range(0,len(newstr_list)):
        if len(newstr_list[i])<=2 or newstr_list[i]==None:
            break
        if newstr_list[i][0:3]=='1. ':
            newstr_list.insert(i,'During '+str(ws[str(chr((col_dict_rem['Hour'])+65))+str(j)].value)+' hour KPI validation, the following issues were observed:')
            break
    ws[nextcol6+str(j)]='\n'.join(newstr_list)

for row in ws.iter_rows():
    for cell in row:
        cell.alignment = Alignment(wrap_text=True,vertical='center')
for col in ws.iter_cols():
    for cell in col:
        cell.alignment = Alignment(wrap_text=True,vertical='center')
        
wb.save(file_name_rem)
wb.close()
    
print ("\n\n\t\t\t\t\t***TASK SUCCESSFULLY COMPLETED***")
print("\n\t\t\t\t\t\t\tPAG, AMDOCS")
print("\n\n\n\n\n\nVersion 3.3 by Vatsal Gupta")
input("\nPress ENTER to Exit...")
